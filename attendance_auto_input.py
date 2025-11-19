#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
勤務データ自動入力ツール
CSVファイルの勤務データをExcelファイルに自動入力します。
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import argparse
from config import EXCEL_CONFIG

class AttendanceAutoInput:
    def __init__(self, csv_file, excel_file, surname=None, full_name=None):
        self.csv_file = csv_file
        self.excel_file = excel_file
        self.surname = surname
        self.full_name = full_name
        self.workbook = None
        self.worksheet = None
        
    def load_csv_data(self):
        """CSVファイルから勤務データを読み込み"""
        try:
            # CSVファイルを読み込み（UTF-8 BOM対応）
            df = pd.read_csv(self.csv_file, encoding='utf-8-sig')
            print(f"CSVファイルを読み込みました: {len(df)}行のデータ")
            return df
        except Exception as e:
            print(f"CSVファイルの読み込みエラー: {e}")
            return None
    
    def load_excel_file(self):
        """Excelファイルを読み込み"""
        try:
            self.workbook = load_workbook(self.excel_file)
            # 最初のシート（業務報告書）を選択
            self.worksheet = self.workbook.worksheets[0]
            print(f"Excelファイルを読み込みました: {self.excel_file}")
            return True
        except Exception as e:
            print(f"Excelファイルの読み込みエラー: {e}")
            return False
    
    def parse_time(self, time_str):
        """時刻文字列をdatetimeオブジェクトに変換"""
        if pd.isna(time_str) or time_str == "":
            return None
        try:
            # "HH:MM"形式の時刻を解析
            return datetime.strptime(str(time_str), "%H:%M").time()
        except:
            return None
    
    def time_to_duration(self, time_str):
        """時間文字列（HH:MM形式）をtimedelta形式に変換"""
        if pd.isna(time_str) or time_str == "":
            return timedelta(0)
        try:
            hours, minutes = map(int, str(time_str).split(':'))
            return timedelta(hours=hours, minutes=minutes)
        except:
            return timedelta(0)
    
    def calculate_work_hours(self, start_time, end_time, break_time):
        """労働時間を計算"""
        if not start_time or not end_time:
            return timedelta(0)
        
        # 開始時刻と終了時刻からdatetimeオブジェクトを作成
        start_dt = datetime.combine(datetime.today(), start_time)
        end_dt = datetime.combine(datetime.today(), end_time)
        
        # 終了時刻が開始時刻より早い場合（日をまたぐ場合）
        if end_dt < start_dt:
            end_dt += timedelta(days=1)
        
        # 労働時間 = 終了時刻 - 開始時刻 - 休憩時間
        work_duration = end_dt - start_dt - break_time
        return work_duration
    
    def update_excel_data(self, csv_data):
        """ExcelファイルにCSVデータを反映"""
        if not self.worksheet:
            print("Excelワークシートが読み込まれていません")
            return False
        
        # データ開始行（12行目がヘッダー、13行目からデータ）
        data_start_row = 13
        
        total_work_time = timedelta(0)
        work_days = 0
        
        for index, row in csv_data.iterrows():
            excel_row = data_start_row + index
            
            # 日付の解析
            date_str = row['日付']
            try:
                date_obj = datetime.strptime(date_str, "%Y/%m/%d")
                # B列: 日付
                self.worksheet.cell(row=excel_row, column=2, value=date_obj)
                
                # C列: 曜日
                weekdays = ['月', '火', '水', '木', '金', '土', '日']
                weekday = weekdays[date_obj.weekday()]
                self.worksheet.cell(row=excel_row, column=3, value=weekday)
                
            except ValueError:
                print(f"日付の解析エラー: {date_str}")
                continue
            
            # 休日区分の確認
            holiday_type = row['休日区分']
            attendance_status = row['勤怠状況']
            
            # 出勤・退勤時刻の処理
            start_time = self.parse_time(row['出勤時刻'])
            end_time = self.parse_time(row['退勤時刻'])
            break_time = self.time_to_duration(row['休憩時間'])
            
            # 休日判定: 公休・法休・祝日、または「有休」かつ「出勤・退勤時刻がない」場合
            is_holiday_type = holiday_type in ['公休', '法休', '祝日']
            has_work_hours = start_time is not None and end_time is not None
            
            if is_holiday_type or (attendance_status == '有休' and not has_work_hours):
                # 休日の場合
                self.worksheet.cell(row=excel_row, column=4, value=None)  # 出勤
                self.worksheet.cell(row=excel_row, column=5, value=None)  # 退勤
                self.worksheet.cell(row=excel_row, column=6, value=None)  # 休憩
                self.worksheet.cell(row=excel_row, column=7, value=timedelta(0))  # 稼働時間
                
                if attendance_status == '有休':
                    self.worksheet.cell(row=excel_row, column=8, value='有休')
                else:
                    self.worksheet.cell(row=excel_row, column=8, value='')
                    
            elif start_time and end_time:
                # 通常勤務の場合
                work_hours = self.calculate_work_hours(start_time, end_time, break_time)
                
                # D列: 出勤時刻
                self.worksheet.cell(row=excel_row, column=4, value=start_time)
                # E列: 退勤時刻  
                self.worksheet.cell(row=excel_row, column=5, value=end_time)
                # F列: 休憩時間
                self.worksheet.cell(row=excel_row, column=6, value=break_time)
                # G列: 稼働時間
                self.worksheet.cell(row=excel_row, column=7, value=work_hours)
                # H列: 作業内容
                # 基本は空欄（ユーザー要望）
                self.worksheet.cell(row=excel_row, column=8, value='')
                
                total_work_time += work_hours
                work_days += 1
            else:
                # データが不完全な場合
                self.worksheet.cell(row=excel_row, column=4, value=None)
                self.worksheet.cell(row=excel_row, column=5, value=None)
                self.worksheet.cell(row=excel_row, column=6, value=None)
                self.worksheet.cell(row=excel_row, column=7, value=timedelta(0))
                self.worksheet.cell(row=excel_row, column=8, value='')
        
        # サマリー情報の更新
        # 出勤日数（9行目、C列）
        self.worksheet.cell(row=9, column=3, value=work_days)
        # 稼働時間（10行目、C列）
        self.worksheet.cell(row=10, column=3, value=total_work_time)
        
        # 氏名の入力（C7）
        if self.full_name:
            name_row = EXCEL_CONFIG['summary'].get('name_cell_row', 7)
            name_col = EXCEL_CONFIG['summary'].get('name_cell_col', 3)
            self.worksheet.cell(row=name_row, column=name_col, value=self.full_name)
            print(f"氏名を入力しました: {self.full_name}")
        
        print(f"データ更新完了: 出勤日数 {work_days}日, 総稼働時間 {total_work_time}")
        return True
    
    def save_excel_file(self, output_file=None):
        """Excelファイルを保存"""
        if not self.workbook:
            print("保存するワークブックがありません")
            return False
        
        try:
            if output_file:
                self.workbook.save(output_file)
                print(f"Excelファイルを保存しました: {output_file}")
                return True
            else:
                # output_fileが指定されていない場合は、BytesIOに保存して返す（Streamlit用）
                from io import BytesIO
                stream = BytesIO()
                self.workbook.save(stream)
                stream.seek(0)
                return stream
        except Exception as e:
            print(f"Excelファイルの保存エラー: {e}")
            return False
    
    def process(self, output_file=None):
        """メイン処理"""
        print("=== 勤務データ自動入力ツール ===")
        
        # CSVデータの読み込み
        csv_data = self.load_csv_data()
        if csv_data is None:
            return False
        
        # Excelファイルの読み込み
        if not self.load_excel_file():
            return False
        
        # データの更新
        if not self.update_excel_data(csv_data):
            return False
        
        # ファイルの保存
        result = self.save_excel_file(output_file)
        if not result:
            return False
        
        print("処理が完了しました！")
        return result

def main():
    parser = argparse.ArgumentParser(description='勤務データ自動入力ツール')
    parser.add_argument('csv_file', help='入力CSVファイルのパス')
    parser.add_argument('excel_file', help='対象Excelファイルのパス')
    parser.add_argument('-o', '--output', help='出力Excelファイルのパス（省略時は元ファイルを上書き）')
    
    args = parser.parse_args()
    
    # ファイルの存在確認
    if not os.path.exists(args.csv_file):
        print(f"CSVファイルが見つかりません: {args.csv_file}")
        return 1
    
    if not os.path.exists(args.excel_file):
        print(f"Excelファイルが見つかりません: {args.excel_file}")
        return 1
    
    # 処理実行
    # コマンドライン引数では名前の入力は未対応（必要なら追加）
    tool = AttendanceAutoInput(args.csv_file, args.excel_file)
    success = tool.process(args.output)
    
    return 0 if success else 1

if __name__ == "__main__":
    sys.exit(main())
